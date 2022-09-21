import csv
import os
import zipfile
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

def test_create_zip():
    zip = zipfile.ZipFile(os.path.abspath('zip.zip'), 'w')
    zip.write('./resources/xlsx.xlsx')
    zip.write('./resources/csv.csv')
    zip.write('./resources/pdf.pdf')
    zip.close()

def test_check_pdf():
    with zipfile.ZipFile(os.path.abspath("./zip.zip")) as zip1:
        with zip1.open('resources/pdf.pdf') as pdfFile:
            pdfFile = PdfReader(pdfFile)
            number_of_pages = len(pdfFile.pages)
            assert number_of_pages == 1

def test_check_csv():
    with zipfile.ZipFile(os.path.abspath("./zip.zip")) as zip2:
        with zip2.open('resources/csv.csv') as csvFile  :
            text = csvFile.read().decode()
            assert '100' in text


def test_check_xlsx():
    with zipfile.ZipFile(os.path.abspath("./zip.zip")) as zip3:
        with zip3.open('resources/xlsx.xlsx') as xlfile:
            file = load_workbook(xlfile)
            text = file.active.cell(row=3, column=2).value
            assert 'Petr' in text